import streamlit as st
import pandas as pd
import openpyxl
import csv
import replicate
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
from io import BytesIO
import requests

# Настройка страницы
st.set_page_config(
    page_title="Генератор учебных заданий",
    page_icon="🎓",
    layout="wide"
)

# Подключаем шрифт Golos Text из Google Fonts
st.markdown("""
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Golos+Text:wght@400;500;600;700&display=swap" rel="stylesheet">
    <link href="https://fonts.googleapis.com/css2?family=Material+Icons&display=swap" rel="stylesheet">

    <style>
        /* Применяем Golos Text ко всем элементам, кроме иконок Material */
        *:not([data-testid="stIconMaterial"]) {
            font-family: 'Golos Text', sans-serif !important;
        }

        /* Полностью скрываем текст keyboard_double_arrow_right */
        [data-testid="collapsedControl"] [data-testid="stIconMaterial"],
        [data-testid="stSidebarCollapsedControl"] [data-testid="stIconMaterial"] {
            visibility: hidden !important;
            position: relative !important;
            width: 40px !important;
            height: 40px !important;
        }

        /* Показываем нашу иконку меню вместо keyboard_double_arrow_right */
        [data-testid="collapsedControl"] [data-testid="stIconMaterial"]::before,
        [data-testid="stSidebarCollapsedControl"] [data-testid="stIconMaterial"]::before {
            content: "☰" !important;
            visibility: visible !important;
            position: absolute !important;
            top: 50% !important;
            left: 50% !important;
            transform: translate(-50%, -50%) !important;
            font-size: 24px !important;
            color: rgba(49, 51, 63, 0.6) !important;
            font-family: Arial, sans-serif !important;
            line-height: 1 !important;
        }

        [data-testid="collapsedControl"]:hover [data-testid="stIconMaterial"]::before,
        [data-testid="stSidebarCollapsedControl"]:hover [data-testid="stIconMaterial"]::before {
            color: rgba(49, 51, 63, 1) !important;
        }
    </style>
""", unsafe_allow_html=True)

# Инициализация session_state
if 'uploaded_file' not in st.session_state:
    st.session_state.uploaded_file = None
if 'test_results' not in st.session_state:
    st.session_state.test_results = None
if 'chosen_model' not in st.session_state:
    st.session_state.chosen_model = None
if 'chosen_program' not in st.session_state:
    st.session_state.chosen_program = None
if 'processed_data' not in st.session_state:
    st.session_state.processed_data = None
if 'show_model_selector' not in st.session_state:
    st.session_state.show_model_selector = False

# Получаем API ключи из secrets
REPLICATE_API_TOKEN = st.secrets.get("REPLICATE_API_TOKEN", "")

# Инициализация клиента Replicate
os.environ["REPLICATE_API_TOKEN"] = REPLICATE_API_TOKEN

# Стоимость за задание для каждой модели (в USD)
MODEL_COSTS = {
    "deepseek": 0.0002,     # DeepSeek-V3: ~$0.0002 за задание
    "claude": 0.005,        # Claude Sonnet 3.5: ~$0.005 за задание
    "gpt4o": 0.004,         # GPT-4o: ~$0.004 за задание
    "llama": 0.002,         # Llama 3.1 405B: ~$0.002 за задание
    "gemini_flash": 0.004,  # Gemini 2.5 Flash: ~$0.004 за задание
    "gpt51": 0.018,         # GPT-5.1 (high): ~$0.018 за задание (с reasoning)
    "kimi": 0.0038          # Kimi K2: ~$0.0038 за задание
}

# Примерное время обработки одной задачи (в секундах)
# При параллельной обработке с 10 потоками
MODEL_TIME_PER_TASK = {
    "deepseek": 4,      # DeepSeek-V3: ~4 сек
    "claude": 5,        # Claude Sonnet 3.5: ~5 сек
    "gpt4o": 4,         # GPT-4o: ~4 сек
    "llama": 7,         # Llama 3.1 405B: ~7 сек
    "gemini_flash": 3,  # Gemini 2.5 Flash: ~3 сек
    "gpt51": 9,         # GPT-5.1 (high): ~9 сек (с reasoning)
    "kimi": 5           # Kimi K2: ~5 сек
}

# ============================================================================
# ФУНКЦИИ ДЛЯ РАБОТЫ С КУРСОМ ВАЛЮТ
# ============================================================================

@st.cache_data(ttl=3600)  # Кэшировать на 1 час
def get_usd_rub_rate():
    """Получает курс USD/RUB из ЦБ РФ"""
    try:
        response = requests.get("https://www.cbr-xml-daily.ru/daily_json.js", timeout=5)
        data = response.json()
        rate = data["Valute"]["USD"]["Value"]
        return rate
    except Exception as e:
        # Если не удалось получить курс, используем примерный курс
        st.warning(f"⚠️ Не удалось получить курс ЦБ РФ, используется примерный курс 90 ₽")
        return 90.0

def calculate_cost(num_tasks, model_key, usd_rub_rate):
    """Рассчитывает стоимость в USD и RUB"""
    cost_usd = num_tasks * MODEL_COSTS[model_key]
    cost_rub = cost_usd * usd_rub_rate
    return cost_usd, cost_rub

def calculate_time(num_tasks, model_key):
    """Рассчитывает примерное время обработки с учётом параллелизма (10 потоков)"""
    time_per_task = MODEL_TIME_PER_TASK[model_key]
    # При 10 параллельных потоках делим на 10, но добавляем небольшой оверхед
    total_seconds = (num_tasks * time_per_task) / 10 * 1.2

    # Форматируем в минуты и секунды
    if total_seconds < 60:
        return f"~{int(total_seconds)} сек"
    else:
        minutes = int(total_seconds // 60)
        seconds = int(total_seconds % 60)
        if seconds > 0:
            return f"~{minutes} мин {seconds} сек"
        else:
            return f"~{minutes} мин"

# ============================================================================
# ФУНКЦИИ ДЛЯ РАБОТЫ С ФАЙЛАМИ
# ============================================================================

def load_prompts():
    """Загружает промпты из promts.csv"""
    try:
        prompts = {}
        with open("promts.csv", 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                level = row['Уровень сложности'].strip()
                prompt_text = row['Промпт'].strip()

                if level and prompt_text:
                    prompts[level] = prompt_text

        return prompts
    except Exception as e:
        st.error(f"Ошибка чтения promts.csv: {e}")
        return {}

def load_excel(file):
    """Загружает Excel файл и возвращает workbook"""
    try:
        wb = openpyxl.load_workbook(file)
        return wb
    except Exception as e:
        st.error(f"Ошибка загрузки Excel: {e}")
        return None

def get_educational_programs(wb):
    """Извлекает уникальные образовательные программы из Excel"""
    ws = wb.active

    # Находим заголовки
    headers = {}
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(1, col).value
        if cell_value:
            headers[cell_value.strip()] = col

    col_program = headers.get('Образовательная программа')

    if not col_program:
        return []

    programs = set()
    for row in range(2, ws.max_row + 1):
        program = ws.cell(row, col_program).value
        if program:
            programs.add(program.strip())

    return sorted(list(programs))

def count_available_tasks_per_program(wb):
    """Подсчитывает количество доступных задач для каждой программы"""
    ws = wb.active

    # Находим заголовки
    headers = {}
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(1, col).value
        if cell_value:
            headers[cell_value.strip()] = col

    col_program = headers.get('Образовательная программа')
    col_discipline = headers.get('Дисциплина / модуль / практика')
    col_level = headers.get('Уровень сложности')
    col_task = headers.get('Задание')

    if not col_program:
        return {}

    program_counts = {}
    prompts = load_prompts()

    for row in range(2, ws.max_row + 1):
        program = ws.cell(row, col_program).value
        discipline = ws.cell(row, col_discipline).value
        level = ws.cell(row, col_level).value
        current_task = ws.cell(row, col_task).value

        # Проверяем что есть промпт для этого уровня сложности
        if program and discipline and level and not current_task:
            prompt_template = prompts.get(level)
            if prompt_template:  # Считаем только если есть промпт
                program = program.strip()
                if program not in program_counts:
                    program_counts[program] = 0
                program_counts[program] += 1

    return program_counts

def count_total_tasks(wb):
    """Подсчитывает общее количество доступных задач во всём файле"""
    ws = wb.active

    # Находим заголовки
    headers = {}
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(1, col).value
        if cell_value:
            headers[cell_value.strip()] = col

    col_discipline = headers.get('Дисциплина / модуль / практика')
    col_level = headers.get('Уровень сложности')
    col_task = headers.get('Задание')

    total_count = 0
    prompts = load_prompts()

    for row in range(2, ws.max_row + 1):
        discipline = ws.cell(row, col_discipline).value
        level = ws.cell(row, col_level).value
        current_task = ws.cell(row, col_task).value

        # Проверяем что есть промпт для этого уровня сложности
        if discipline and level and not current_task:
            prompt_template = prompts.get(level)
            if prompt_template:  # Считаем только если есть промпт
                total_count += 1

    return total_count

def get_tasks_from_excel(wb, max_rows=None, filter_program=None):
    """Извлекает задачи из Excel"""
    ws = wb.active

    # Находим заголовки
    headers = {}
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(1, col).value
        if cell_value:
            headers[cell_value.strip()] = col

    col_program = headers.get('Образовательная программа')
    col_discipline = headers.get('Дисциплина / модуль / практика')
    col_level = headers.get('Уровень сложности')
    col_task = headers.get('Задание')
    col_answer = headers.get('Ключ (ответ)')

    tasks = []
    prompts = load_prompts()

    # Если фильтруем по программе, сначала собираем ВСЕ задачи этой программы,
    # а потом ограничиваем количество. Иначе можем пропустить строки программы.
    if filter_program:
        # Сканируем весь файл для поиска строк нужной программы
        for row in range(2, ws.max_row + 1):
            program = ws.cell(row, col_program).value if col_program else None
            discipline = ws.cell(row, col_discipline).value
            level = ws.cell(row, col_level).value
            current_task = ws.cell(row, col_task).value

            # Нормализуем строки для сравнения (убираем пробелы по краям)
            program_normalized = program.strip() if program else None
            filter_normalized = filter_program.strip() if filter_program else None

            # Фильтруем по программе
            if program_normalized != filter_normalized:
                continue

            if discipline and level and not current_task:
                prompt_template = prompts.get(level)
                if prompt_template:
                    tasks.append({
                        'row': row,
                        'program': program,
                        'discipline': discipline,
                        'level': level,
                        'prompt': prompt_template
                    })

                    # Ограничиваем количество задач после сбора
                    if max_rows and len(tasks) >= max_rows:
                        break
    else:
        # Без фильтра по программе - используем старую логику
        row_limit = min(max_rows + 2, ws.max_row + 1) if max_rows else ws.max_row + 1

        for row in range(2, row_limit):
            program = ws.cell(row, col_program).value if col_program else None
            discipline = ws.cell(row, col_discipline).value
            level = ws.cell(row, col_level).value
            current_task = ws.cell(row, col_task).value

            if discipline and level and not current_task:
                prompt_template = prompts.get(level)
                if prompt_template:
                    tasks.append({
                        'row': row,
                        'program': program,
                        'discipline': discipline,
                        'level': level,
                        'prompt': prompt_template
                    })

    return tasks, (col_task, col_answer)

# ============================================================================
# ФУНКЦИИ ГЕНЕРАЦИИ ДЛЯ РАЗНЫХ МОДЕЛЕЙ
# ============================================================================

def generate_deepseek(discipline, level, prompt_template):
    """Генерация через DeepSeek-V3"""
    full_prompt = f"""{prompt_template}

Дисциплина/модуль/практика: {discipline}

Сгенерируй задание и ответ к нему в следующем формате:

ЗАДАНИЕ:
[текст задания]

КЛЮЧ (ОТВЕТ):
[правильный ответ]

Важно: отвечай только на русском языке."""
    
    try:
        output = replicate.run(
            "deepseek-ai/deepseek-v3",
            input={"prompt": full_prompt, "max_tokens": 2000, "temperature": 0.7}
        )
        
        response_text = ""
        for item in output:
            response_text += item
        
        return parse_response(response_text)
    except Exception as e:
        return None, None, str(e)

def generate_claude(discipline, level, prompt_template):
    """Генерация через Claude Sonnet 3.5 (via Replicate)"""
    full_prompt = f"""{prompt_template}

Дисциплина/модуль/практика: {discipline}

Сгенерируй задание и ответ к нему в следующем формате:

ЗАДАНИЕ:
[текст задания]

КЛЮЧ (ОТВЕТ):
[правильный ответ]

Важно: отвечай только на русском языке."""
    
    try:
        output = replicate.run(
            "anthropic/claude-3.5-sonnet",
            input={
                "prompt": full_prompt,
                "max_tokens": 2000,
                "temperature": 0.7
            }
        )
        
        response_text = ""
        for item in output:
            response_text += item
        
        return parse_response(response_text)
    except Exception as e:
        return None, None, str(e)

def generate_gpt4o(discipline, level, prompt_template):
    """Генерация через GPT-4o (via Replicate)"""
    full_prompt = f"""{prompt_template}

Дисциплина/модуль/практика: {discipline}

Сгенерируй задание и ответ к нему в следующем формате:

ЗАДАНИЕ:
[текст задания]

КЛЮЧ (ОТВЕТ):
[правильный ответ]

Важно: отвечай только на русском языке."""

    try:
        output = replicate.run(
            "openai/gpt-4o",
            input={
                "prompt": full_prompt,
                "max_tokens": 2000,
                "temperature": 0.7
            }
        )

        response_text = ""
        for item in output:
            response_text += item

        return parse_response(response_text)
    except Exception as e:
        return None, None, str(e)

def generate_llama(discipline, level, prompt_template):
    """Генерация через Llama 3.1 405B Instruct (via Replicate)"""
    full_prompt = f"""{prompt_template}

Дисциплина/модуль/практика: {discipline}

Сгенерируй задание и ответ к нему в следующем формате:

ЗАДАНИЕ:
[текст задания]

КЛЮЧ (ОТВЕТ):
[правильный ответ]

Важно: отвечай только на русском языке."""

    try:
        output = replicate.run(
            "meta/meta-llama-3.1-405b-instruct",
            input={
                "prompt": full_prompt,
                "max_tokens": 2000,
                "temperature": 0.7
            }
        )

        response_text = ""
        for item in output:
            response_text += item

        return parse_response(response_text)
    except Exception as e:
        return None, None, str(e)

def generate_gemini_flash(discipline, level, prompt_template):
    """Генерация через Gemini 2.5 Flash (via Replicate)"""
    full_prompt = f"""{prompt_template}

Дисциплина/модуль/практика: {discipline}

Сгенерируй задание и ответ к нему в следующем формате:

ЗАДАНИЕ:
[текст задания]

КЛЮЧ (ОТВЕТ):
[правильный ответ]

Важно: отвечай только на русском языке."""

    try:
        output = replicate.run(
            "google/gemini-2.5-flash",
            input={
                "prompt": full_prompt,
                "max_tokens": 2000,
                "temperature": 0.7
            }
        )

        response_text = ""
        for item in output:
            response_text += item

        return parse_response(response_text)
    except Exception as e:
        return None, None, str(e)

def generate_gpt51(discipline, level, prompt_template):
    """Генерация через GPT-5.1 (high reasoning) (via Replicate)"""
    full_prompt = f"""{prompt_template}

Дисциплина/модуль/практика: {discipline}

Сгенерируй задание и ответ к нему в следующем формате:

ЗАДАНИЕ:
[текст задания]

КЛЮЧ (ОТВЕТ):
[правильный ответ]

Важно: отвечай только на русском языке."""

    try:
        output = replicate.run(
            "openai/gpt-5.1",
            input={
                "prompt": full_prompt,
                "max_tokens": 2000,
                "temperature": 0.7,
                "reasoning_effort": "high"
            }
        )

        response_text = ""
        for item in output:
            response_text += item

        return parse_response(response_text)
    except Exception as e:
        return None, None, str(e)

def generate_kimi(discipline, level, prompt_template):
    """Генерация через Kimi K2 (via Replicate)"""
    full_prompt = f"""{prompt_template}

Дисциплина/модуль/практика: {discipline}

Сгенерируй задание и ответ к нему в следующем формате:

ЗАДАНИЕ:
[текст задания]

КЛЮЧ (ОТВЕТ):
[правильный ответ]

Важно: отвечай только на русском языке."""

    try:
        output = replicate.run(
            "moonshotai/kimi-k2-instruct",
            input={
                "prompt": full_prompt,
                "max_tokens": 2000,
                "temperature": 0.7
            }
        )

        response_text = ""
        for item in output:
            response_text += item

        return parse_response(response_text)
    except Exception as e:
        return None, None, str(e)

def parse_response(response_text):
    """Парсит ответ модели"""
    task = ""
    answer = ""
    
    if "ЗАДАНИЕ:" in response_text and "КЛЮЧ (ОТВЕТ):" in response_text:
        parts = response_text.split("КЛЮЧ (ОТВЕТ):")
        task = parts[0].replace("ЗАДАНИЕ:", "").strip()
        answer = parts[1].strip()
    elif "ЗАДАНИЕ:" in response_text and "ОТВЕТ:" in response_text:
        parts = response_text.split("ОТВЕТ:")
        task = parts[0].replace("ЗАДАНИЕ:", "").strip()
        answer = parts[1].strip()
    else:
        lines = response_text.strip().split('\n')
        mid = len(lines) // 2
        task = '\n'.join(lines[:mid]).strip()
        answer = '\n'.join(lines[mid:]).strip()
    
    return task, answer, None

# ============================================================================
# ОСНОВНОЕ ПРИЛОЖЕНИЕ
# ============================================================================

# Боковое меню со справкой
with st.sidebar:
    # Добавляем логотип БФУ
    st.image("bfu.png", use_container_width=True)
    st.markdown("---")

    st.markdown("## 📖 Как пользоваться")
    st.markdown("""
    ### Быстрый старт

    **1. Загрузите файл**
    Выберите Excel файл с заданиями

    **2. Тестируйте модели**
    Посмотрите примеры работы 7 AI моделей

    **3. Выберите модель**
    Определитесь с лучшим вариантом по цене/качеству

    **4. Выберите программу**
    Укажите образовательную программу для генерации

    **5. Настройте объём**
    Выберите количество заданий (до 2000 строк)

    **6. Запустите обработку**
    AI сгенерирует задания и ответы

    **7. Скачайте результат**
    Получите готовый Excel файл

    ---

    ### 💡 Полезно знать

    - **Параллельная обработка** — 10 заданий одновременно
    - **Точная стоимость** — расчёт по курсу ЦБ РФ
    - **Время обработки** — ~3-9 сек на задание
    - **7 моделей** — DeepSeek, Gemini, GPT, Claude, Llama, Kimi

    ---

    *Генератор учебных заданий v1.0*
    """)

st.title("🎓 Генератор учебных заданий")
st.markdown("Автоматическая генерация заданий через AI модели")

# Шаг 1: Загрузка файла
st.header("1️⃣ Загрузите файл")
uploaded_file = st.file_uploader("Выберите megaphops.xlsx", type=['xlsx'])

if uploaded_file:
    st.session_state.uploaded_file = uploaded_file
    st.success(f"✅ Файл загружен: {uploaded_file.name}")

    # Две кнопки: показать варианты и выбрать модель сразу
    col1, col2 = st.columns(2)

    with col1:
        show_variants = st.button("🔍 Показать варианты заданий", type="primary", use_container_width=True)

    with col2:
        select_model = st.button("⚡ Выбрать модель сразу", type="secondary", use_container_width=True)

    if select_model:
        st.session_state.show_model_selector = True

    if show_variants:
        with st.spinner("Тестируем 7 AI моделей на первых 2 заданиях..."):
            wb = load_excel(uploaded_file)
            if wb:
                tasks, cols = get_tasks_from_excel(wb, max_rows=2)

                if len(tasks) >= 2:
                    results = {
                        "DeepSeek-V3": [],
                        "Claude Sonnet 3.5": [],
                        "GPT-4o": [],
                        "Llama 3.1 405B": [],
                        "Gemini 2.5 Flash": [],
                        "GPT-5.1 (high)": [],
                        "Kimi K2": []
                    }

                    # Генерируем для первых 2 заданий
                    for task in tasks[:2]:
                        # DeepSeek
                        task_text, answer_text, error = generate_deepseek(
                            task['discipline'], task['level'], task['prompt']
                        )
                        results["DeepSeek-V3"].append({
                            "Дисциплина": task['discipline'],
                            "Задание": task_text if task_text else f"Ошибка: {error}",
                            "Ответ": answer_text if answer_text else ""
                        })

                        # Claude
                        task_text, answer_text, error = generate_claude(
                            task['discipline'], task['level'], task['prompt']
                        )
                        results["Claude Sonnet 3.5"].append({
                            "Дисциплина": task['discipline'],
                            "Задание": task_text if task_text else f"Ошибка: {error}",
                            "Ответ": answer_text if answer_text else ""
                        })

                        # GPT-4o
                        task_text, answer_text, error = generate_gpt4o(
                            task['discipline'], task['level'], task['prompt']
                        )
                        results["GPT-4o"].append({
                            "Дисциплина": task['discipline'],
                            "Задание": task_text if task_text else f"Ошибка: {error}",
                            "Ответ": answer_text if answer_text else ""
                        })

                        # Llama 3.1 405B
                        task_text, answer_text, error = generate_llama(
                            task['discipline'], task['level'], task['prompt']
                        )
                        results["Llama 3.1 405B"].append({
                            "Дисциплина": task['discipline'],
                            "Задание": task_text if task_text else f"Ошибка: {error}",
                            "Ответ": answer_text if answer_text else ""
                        })

                        # Gemini 2.5 Flash
                        task_text, answer_text, error = generate_gemini_flash(
                            task['discipline'], task['level'], task['prompt']
                        )
                        results["Gemini 2.5 Flash"].append({
                            "Дисциплина": task['discipline'],
                            "Задание": task_text if task_text else f"Ошибка: {error}",
                            "Ответ": answer_text if answer_text else ""
                        })

                        # GPT-5.1 (high)
                        task_text, answer_text, error = generate_gpt51(
                            task['discipline'], task['level'], task['prompt']
                        )
                        results["GPT-5.1 (high)"].append({
                            "Дисциплина": task['discipline'],
                            "Задание": task_text if task_text else f"Ошибка: {error}",
                            "Ответ": answer_text if answer_text else ""
                        })

                        # Kimi K2
                        task_text, answer_text, error = generate_kimi(
                            task['discipline'], task['level'], task['prompt']
                        )
                        results["Kimi K2"].append({
                            "Дисциплина": task['discipline'],
                            "Задание": task_text if task_text else f"Ошибка: {error}",
                            "Ответ": answer_text if answer_text else ""
                        })

                    st.session_state.test_results = results
                else:
                    st.error("В файле недостаточно пустых строк для тестирования")

# Прямой выбор модели без тестирования
if st.session_state.show_model_selector and not st.session_state.chosen_model:
    st.header("2️⃣ Выберите модель")
    st.markdown("Выберите AI модель для генерации заданий:")

    # Получаем курс доллара для расчетов
    usd_rub_rate = get_usd_rub_rate()

    # Получаем общее количество задач в файле
    wb = load_excel(st.session_state.uploaded_file)
    total_tasks_count = count_total_tasks(wb) if wb else 0

    models = {
        "DeepSeek-V3": {
            "icon": "🚀",
            "description": "Лучшая цена/качество. $0.14 за 1M токенов",
            "key": "deepseek"
        },
        "Claude Sonnet 3.5": {
            "icon": "🧠",
            "description": "Топовое качество. $3 за 1M токenов",
            "key": "claude"
        },
        "GPT-4o": {
            "icon": "⚡",
            "description": "Быстрый и качественный. $2.50 за 1M токенов",
            "key": "gpt4o"
        },
        "Llama 3.1 405B": {
            "icon": "🦙",
            "description": "405B параметров. Мощная модель",
            "key": "llama"
        },
        "Gemini 2.5 Flash": {
            "icon": "💎",
            "description": "Google быстрый. $0.30 (ввод) + $2.50 (вывод) за 1M",
            "key": "gemini_flash"
        },
        "GPT-5.1 (high)": {
            "icon": "🧪",
            "description": "Топ reasoning. $1.25 (ввод) + $10 (вывод) за 1M",
            "key": "gpt51"
        },
        "Kimi K2": {
            "icon": "🌙",
            "description": "Отлично с русским. $0.15 (ввод) + $2.50 (вывод) за 1M",
            "key": "kimi"
        }
    }

    # Создаем кнопки для выбора модели
    for model_name, model_info in models.items():
        # Рассчитываем стоимость для всей таблицы
        if total_tasks_count > 0:
            full_cost_usd, full_cost_rub = calculate_cost(total_tasks_count, model_info['key'], usd_rub_rate)
            full_time = calculate_time(total_tasks_count, model_info['key'])

            col1, col2 = st.columns([4, 1])
            with col1:
                st.markdown(
                    f"**{model_info['icon']} {model_name}** — {model_info['description']}\n\n"
                    f"💰 {full_cost_rub:.2f} ₽ (${full_cost_usd:.2f}) • ⏱️ {full_time} ({total_tasks_count} задач)"
                )
            with col2:
                if st.button("✅ Выбрать", key=f"direct_choose_{model_info['key']}"):
                    st.session_state.chosen_model = model_info['key']
                    st.session_state.show_model_selector = False
                    st.success(f"Выбрана модель: {model_name}")
                    st.rerun()

        st.markdown("---")

# Шаг 2: Показ результатов тестирования
if st.session_state.test_results:
    st.header("2️⃣ Выберите модель")
    st.markdown("Ниже представлены результаты генерации от 7 моделей:")

    # Получаем курс доллара для расчетов
    usd_rub_rate = get_usd_rub_rate()

    # Получаем общее количество задач в файле
    wb = load_excel(st.session_state.uploaded_file)
    total_tasks_count = count_total_tasks(wb) if wb else 0

    models = {
        "DeepSeek-V3": {
            "icon": "🚀",
            "description": "Лучшая цена/качество. $0.14 за 1M токенов",
            "key": "deepseek"
        },
        "Claude Sonnet 3.5": {
            "icon": "🧠",
            "description": "Топовое качество. $3 за 1M токenов",
            "key": "claude"
        },
        "GPT-4o": {
            "icon": "⚡",
            "description": "Быстрый и качественный. $2.50 за 1M токенов",
            "key": "gpt4o"
        },
        "Llama 3.1 405B": {
            "icon": "🦙",
            "description": "405B параметров. Мощная модель",
            "key": "llama"
        },
        "Gemini 2.5 Flash": {
            "icon": "💎",
            "description": "Google быстрый. $0.30 (ввод) + $2.50 (вывод) за 1M",
            "key": "gemini_flash"
        },
        "GPT-5.1 (high)": {
            "icon": "🧪",
            "description": "Топ reasoning. $1.25 (ввод) + $10 (вывод) за 1M",
            "key": "gpt51"
        },
        "Kimi K2": {
            "icon": "🌙",
            "description": "Отлично с русским. $0.15 (ввод) + $2.50 (вывод) за 1M",
            "key": "kimi"
        }
    }

    for model_name, model_info in models.items():
        with st.expander(f"{model_info['icon']} {model_name} - {model_info['description']}", expanded=False):
            df = pd.DataFrame(st.session_state.test_results[model_name])
            st.dataframe(df, use_container_width=True, hide_index=True)

            if st.button(f"✅ Выбрать", key=f"choose_{model_info['key']}"):
                st.session_state.chosen_model = model_info['key']
                st.success(f"Выбрана модель: {model_name}")
                st.rerun()

    # Показываем стоимость для всей таблицы для каждой модели (под превью)
    if total_tasks_count > 0:
        st.markdown("---")
        st.markdown("#### 💰 Стоимость обработки всей таблицы:")
        for model_name, model_info in models.items():
            full_cost_usd, full_cost_rub = calculate_cost(total_tasks_count, model_info['key'], usd_rub_rate)
            full_time = calculate_time(total_tasks_count, model_info['key'])

            st.markdown(
                f"**{model_info['icon']} {model_name} ({total_tasks_count} задач):** "
                f"💰 {full_cost_rub:.2f} ₽ (${full_cost_usd:.2f}) • "
                f"⏱️ {full_time}"
            )

# Шаг 3: Выбор образовательной программы
if st.session_state.chosen_model and st.session_state.uploaded_file:
    st.header("3️⃣ Выберите образовательную программу")

    wb = load_excel(st.session_state.uploaded_file)
    if wb:
        programs = get_educational_programs(wb)
        program_counts = count_available_tasks_per_program(wb)
        total_tasks = count_total_tasks(wb)

        if programs:
            # Показываем общую статистику
            st.info(f"📊 **Всего доступных строк в файле:** {total_tasks}")

            st.markdown("Доступные образовательные программы:")

            # Создаём кнопки для каждой программы
            cols = st.columns(min(3, len(programs)))  # Максимум 3 колонки
            for idx, program in enumerate(programs):
                col_idx = idx % 3
                available_count = program_counts.get(program, 0)
                with cols[col_idx]:
                    if st.button(
                        f"📚 {program}\n\n({available_count} строк)",
                        key=f"program_{idx}",
                        use_container_width=True
                    ):
                        st.session_state.chosen_program = program
                        st.success(f"Выбрана программа: {program} ({available_count} доступных строк)")
                        st.rerun()
        else:
            st.warning("⚠️ В файле не найдены образовательные программы. Убедитесь, что столбец 'Образовательная программа' заполнен.")

# Шаг 4: Основная обработка
if st.session_state.chosen_model and st.session_state.chosen_program:
    st.header("4️⃣ Обработка заданий")

    # Показываем выбранные параметры
    model_names = {
        "deepseek": "DeepSeek-V3",
        "claude": "Claude Sonnet 3.5",
        "gpt4o": "GPT-4o",
        "llama": "Llama 3.1 405B",
        "gemini_flash": "Gemini 2.5 Flash",
        "gpt51": "GPT-5.1 (high)",
        "kimi": "Kimi K2"
    }

    st.info(
        f"**Выбранная модель:** {model_names[st.session_state.chosen_model]}\n\n"
        f"**Образовательная программа:** {st.session_state.chosen_program}"
    )

    # Выбор количества строк
    batch_size = st.slider(
        "Количество строк для обработки (макс 2000)",
        min_value=10,
        max_value=2000,
        value=100,
        step=10
    )

    # Получаем курс доллара
    usd_rub_rate = get_usd_rub_rate()

    # Рассчитываем примерную стоимость
    estimated_cost_usd, estimated_cost_rub = calculate_cost(
        batch_size,
        st.session_state.chosen_model,
        usd_rub_rate
    )

    # Рассчитываем примерное время обработки
    estimated_time = calculate_time(batch_size, st.session_state.chosen_model)

    # Отображаем стоимость и время
    st.info(
        f"💰 **Примерная стоимость обработки {batch_size} заданий через {model_names[st.session_state.chosen_model]}:**\n\n"
        f"- ${estimated_cost_usd:.4f} USD\n"
        f"- {estimated_cost_rub:.2f} ₽ (курс ЦБ РФ: {usd_rub_rate:.2f} ₽/$)\n\n"
        f"⏱️ **Примерное время обработки:** {estimated_time}"
    )

    if st.button("🚀 Начать обработку", type="primary"):
        with st.spinner(f"Обработка {batch_size} строк программы '{st.session_state.chosen_program}'..."):
            wb = load_excel(st.session_state.uploaded_file)
            if wb:
                tasks, (col_task, col_answer) = get_tasks_from_excel(
                    wb,
                    max_rows=batch_size,
                    filter_program=st.session_state.chosen_program
                )
                ws = wb.active
                
                if tasks:
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    # Выбор функции генерации
                    if st.session_state.chosen_model == "deepseek":
                        generate_func = generate_deepseek
                    elif st.session_state.chosen_model == "claude":
                        generate_func = generate_claude
                    elif st.session_state.chosen_model == "gpt4o":
                        generate_func = generate_gpt4o
                    elif st.session_state.chosen_model == "llama":
                        generate_func = generate_llama
                    elif st.session_state.chosen_model == "gemini_flash":
                        generate_func = generate_gemini_flash
                    elif st.session_state.chosen_model == "gpt51":
                        generate_func = generate_gpt51
                    elif st.session_state.chosen_model == "kimi":
                        generate_func = generate_kimi
                    else:
                        generate_func = generate_deepseek  # Fallback
                    
                    results = []
                    errors = 0
                    
                    # Параллельная обработка с 10 потоками
                    MAX_WORKERS = 10
                    
                    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
                        futures = {
                            executor.submit(
                                generate_func,
                                task['discipline'],
                                task['level'],
                                task['prompt']
                            ): task for task in tasks
                        }
                        
                        completed = 0
                        for future in as_completed(futures):
                            task = futures[future]
                            try:
                                task_text, answer_text, error = future.result()
                                
                                if task_text and answer_text:
                                    ws.cell(task['row'], col_task, task_text)
                                    ws.cell(task['row'], col_answer, answer_text)
                                    results.append({
                                        "Строка": task['row'],
                                        "Дисциплина": task['discipline'],
                                        "Задание": task_text[:100] + "...",
                                        "Ответ": answer_text[:100] + "..."
                                    })
                                else:
                                    errors += 1
                            except Exception as e:
                                errors += 1
                            
                            completed += 1
                            progress = completed / len(tasks)
                            progress_bar.progress(progress)
                            status_text.text(f"Обработано: {completed}/{len(tasks)}")
                    
                    # Сохраняем результат
                    output = BytesIO()
                    wb.save(output)
                    output.seek(0)
                    st.session_state.processed_data = output

                    # Рассчитываем фактическую стоимость
                    actual_cost_usd, actual_cost_rub = calculate_cost(
                        len(results),
                        st.session_state.chosen_model,
                        usd_rub_rate
                    )

                    st.success(f"✅ Обработка завершена! Успешно: {len(results)}, Ошибок: {errors}")

                    # Показываем фактическую стоимость
                    st.info(
                        f"💳 **Фактическая стоимость обработки {len(results)} заданий:**\n\n"
                        f"- ${actual_cost_usd:.4f} USD\n"
                        f"- {actual_cost_rub:.2f} ₽"
                    )

                    # Превью результатов
                    st.subheader("📊 Превью результатов")
                    df_results = pd.DataFrame(results)
                    st.dataframe(df_results, width='stretch')
                else:
                    st.warning("Нет задач для обработки")

# Шаг 5: Скачивание
if st.session_state.processed_data:
    st.header("5️⃣ Скачать результат")
    st.download_button(
        label="📥 Скачать megaphops_filled.xlsx",
        data=st.session_state.processed_data,
        file_name="megaphops_filled.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary"
    )
